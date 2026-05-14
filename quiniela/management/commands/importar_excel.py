from django.core.management.base import BaseCommand

from quiniela.models import (

    Partido,

    Jornada

)

from openpyxl import load_workbook

from datetime import datetime

import os


class Command(BaseCommand):

    help = 'Importa partidos desde Excel'


    def handle(self, *args, **kwargs):

        ruta_excel = os.path.join(

            'data',

            'mundial2026.xlsx'

        )

        workbook = load_workbook(

            ruta_excel

        )

        sheet = workbook.active

        contador = 0

        for row in sheet.iter_rows(

            min_row=2,

            values_only=True

        ):

            if not row[0]:

                continue

            (

                local,

                visitante,

                grupo,

                jornada_numero,

                fecha_partido,

                hora_partido,

                estadio,

                ciudad,

                pais_sede,

                resultado_real

            ) = row

            if not jornada_numero:

                self.stdout.write(

                    self.style.WARNING(

                        f'Fila ignorada por jornada vacía: {local} vs {visitante}'

                    )

                )

                continue

            jornada, _ = Jornada.objects.get_or_create(

                numero=int(jornada_numero)

            )

            if isinstance(

                fecha_partido,

                datetime

            ):

                fecha_partido = fecha_partido.date()

            if isinstance(

                hora_partido,

                datetime

            ):

                hora_partido = hora_partido.time()

            Partido.objects.update_or_create(

                local=local,

                visitante=visitante,

                defaults={

                    'grupo': grupo,

                    'jornada': jornada,

                    'fecha_partido': fecha_partido,

                    'hora_partido': hora_partido,

                    'estadio': estadio,

                    'ciudad': ciudad,

                    'pais_sede': pais_sede,

                    'resultado_real': resultado_real

                }

            )

            contador += 1

        self.stdout.write(

            self.style.SUCCESS(

                f'{contador} partidos importados correctamente'

            )

        )